import os
import csv
import logging
import re
from pathlib import Path
from logger_utils import log_step



def extract_free_energy(log_file):
    """Parses a Vina log file and returns the top free energy (affinity) score."""
    try:
        with open(log_file, 'r') as f:
            lines = f.readlines()
            
        for i, line in enumerate(lines):
            if "mode |   affinity" in line:
                # Top score is typically 3 lines after the header
                # mode |   affinity | dist from best mode
                #      | (kcal/mol) | rmsd l.b.| rmsd u.b.
                # -----+------------+----------+----------
                #    1        -12.9          0          0
                if i + 3 < len(lines):
                    data_line = lines[i + 3]
                    parts = data_line.split()
                    if parts and parts[0] == "1":
                        return float(parts[1])
    except Exception as e:
        logging.error(f"Error reading {log_file}: {e}")
    return None

def rank_complexes(output_dir, known_ligands=None):
    """Finds all log files in output_dir, extracts energies, and returns a sorted list."""
    output_path = Path(output_dir)
    log_files = list(output_path.rglob("*.log"))
    
    results = []
    for log_file in log_files:
        energy = extract_free_energy(log_file)
        if energy is not None:
            # Complex name can be inferred from log file name (e.g., 1IEP_imatinib_vina.log -> 1IEP_imatinib)
            complex_name = log_file.stem
            if complex_name.endswith("_vina"):
                complex_name = complex_name[:-5]
                
            run_index = "1"
            for part in log_file.parts:
                run_match = re.match(r'^run_(\d+)$', part)
                if run_match:
                    run_index = run_match.group(1)
                    break
            
            protein_pocket = complex_name
            ligand = "N/A"

            # If we know the ligand names, use them to split accurately from the end
            if known_ligands:
                # Sort ligands by length descending to match longest possible ligand name first
                for lig in sorted(known_ligands, key=len, reverse=True):
                    if complex_name.endswith("_" + lig):
                        ligand = lig
                        protein_pocket = complex_name[:-len(lig)-1]
                        break
            
            # Fallback if no known_ligands or no match
            if ligand == "N/A":
                # Try to extract pocket ID first to see where the protein ends
                match = re.match(r'^(.*?)_pocket_(\d+)_(.*)$', complex_name)
                if match:
                    protein = match.group(1)
                    pocket = match.group(2)
                    ligand = match.group(3)
                    results.append((protein, pocket, ligand, run_index, energy))
                    continue
                else:
                    # Fallback: split by first underscore
                    parts = complex_name.split("_", 1)
                    if len(parts) == 2:
                        protein, ligand = parts
                    else:
                        protein, ligand = complex_name, "N/A"
                    pocket = "N/A"
                    results.append((protein, pocket, ligand, run_index, energy))
                    continue

            # If we successfully extracted protein_pocket and ligand using known_ligands
            match = re.match(r'^(.*?)_pocket_(\d+)$', protein_pocket)
            if match:
                protein = match.group(1)
                pocket = match.group(2)
            else:
                protein = protein_pocket
                pocket = "N/A"
                
            results.append((protein, pocket, ligand, run_index, energy))
            
    # Sort by free energy (lowest/most negative first)
    results.sort(key=lambda x: x[4])
    return results

def _get_base_ligand(ligand):
    """Extracts the base ligand name, ignoring isomer suffixes."""
    return ligand.split("_isomer_")[0] if "_isomer_" in ligand else ligand

def _load_mapping_from_csv(csv_path=None):
    """Loads a mapping from the first two columns of a provided CSV file."""
    import csv
    from pathlib import Path
    import logging
    
    mapping = {}
    if not csv_path:
        return mapping
        
    path_obj = Path(csv_path)
    if path_obj.exists():
        try:
            with open(path_obj, 'r', encoding='utf-8') as f:
                first_line = f.readline()
                f.seek(0)
                delimiter = '\t' if '\t' in first_line else ','
                reader = csv.reader(f, delimiter=delimiter)
                for row in reader:
                    if len(row) >= 2:
                        key = row[0].strip().lower()
                        val = row[1].strip()
                        mapping[key] = val
        except Exception as e:
            logging.warning(f"Failed to read CSV mapping {csv_path}: {e}")
    return mapping

def _get_protein_pockets(curated_results):
    """Returns a dictionary mapping protein to a set of its unique pockets."""
    protein_pockets = {}
    for row in curated_results:
        protein, pocket = row[:2]
        if protein not in protein_pockets:
            protein_pockets[protein] = set()
        if pocket != "N/A":
            protein_pockets[protein].add(pocket)
    return protein_pockets

def _format_display_protein_pocket(protein, pocket, protein_to_name, protein_pockets):
    """Formats the protein display name, only appending pocket info if the protein has multiple pockets."""
    display_protein = protein_to_name.get(protein.lower(), protein)
    return f"{display_protein}_pocket_{pocket}" if pocket != "N/A" and len(protein_pockets.get(protein, set())) > 1 else display_protein

from contextlib import contextmanager

@contextmanager
def _prepare_prolif_protein(protein_cif):
    """Context manager to convert CIF to PDB and yield a ProLIF Molecule."""
    import tempfile
    import os
    from pathlib import Path
    import MDAnalysis as mda
    import prolif as plf
    
    with tempfile.NamedTemporaryFile(suffix=".pdb", delete=False) as tmp_pdb:
        tmp_prot_pdb = Path(tmp_pdb.name)
        
    try:
        _convert_to_pdb(protein_cif, "cif", tmp_prot_pdb)
        u = mda.Universe(str(tmp_prot_pdb))
        protein_mol = plf.Molecule.from_mda(u)
        yield protein_mol
    finally:
        try:
            os.unlink(tmp_prot_pdb)
        except OSError:
            pass

def _generate_network_plot(fp, pose, driver, out_png):
    """Generates a 2D ProLIF interaction network and saves a screenshot using Selenium."""
    import tempfile
    import os
    import time
    import logging
    
    view = fp.plot_lignetwork(pose, kind="frame", frame=0)
    
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as tmp_html:
        tmp_html_path = tmp_html.name
        
    try:
        if hasattr(view, 'save'):
            view.save(tmp_html_path)
        elif hasattr(view, 'write_html'):
            view.write_html(tmp_html_path)
        elif hasattr(view, 'data'):
            with open(tmp_html_path, 'w', encoding='utf-8') as f:
                f.write(view.data)
        else:
            with open(tmp_html_path, 'w', encoding='utf-8') as f:
                f.write(str(view))
        
        driver.get(f"file://{tmp_html_path}")
        time.sleep(2)
        driver.save_screenshot(str(out_png))
        logging.debug(f"Saved ProLif visualization: {out_png}")
    finally:
        try:
            os.unlink(tmp_html_path)
        except OSError:
            pass

def _generate_barcode_plot(df_combined, tiff_path, protein_pocket_base):
    """Generates a Barcode plot from a combined DataFrame."""
    import matplotlib.pyplot as plt
    from prolif.plotting.barcode import Barcode
    from logger_utils import log_step
    
    sorted_columns = df_combined.sum().sort_values(ascending=False).index
    df_sorted = df_combined[sorted_columns]

    ax = Barcode(df_sorted).display(
        figsize=(10, min(8, max(4, len(sorted_columns) * 0.3))),
        dpi=300,
        xlabel=""
    )
    
    fig = ax.figure
    plt.setp(ax.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor', fontsize=8)
    fig.savefig(str(tiff_path), format="tiff", dpi=600, bbox_inches="tight")
    plt.close(fig)
    
    log_step(None, f"Saved barcode for {protein_pocket_base}", color="white")

def _generate_heatmap_plot(df_combined, tiff_path, protein_pocket_base):
    """Generates a Heatmap plot from a combined DataFrame."""
    import matplotlib.pyplot as plt
    import seaborn as sns
    from logger_utils import log_step
    
    if df_combined.empty:
        log_step(None, f"No data to plot heatmap for {protein_pocket_base}", color="yellow")
        return

    # Group by interaction type only
    if df_combined.columns.nlevels == 3:
        # Level 2 is the interaction type
        df_grouped = df_combined.groupby(level=2, axis=1).sum()
    elif df_combined.columns.nlevels == 2:
        # Level 1 is the interaction type
        df_grouped = df_combined.groupby(level=1, axis=1).sum()
    else:
        df_grouped = df_combined
        
    df_plot = df_grouped.T
    
    sorted_idx = df_plot.sum(axis=1).sort_values(ascending=False).index
    df_plot = df_plot.loc[sorted_idx]
    
    fig_width = max(10, len(df_plot.columns) * 0.8)
    fig_height = max(6, len(df_plot.index) * 0.8)
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))
    
    sns.heatmap(df_plot.astype(float), cmap='YlGnBu', annot=True, fmt='g', 
                cbar_kws={'label': 'Interaction Count'}, ax=ax)
    
    plt.title(f"Interaction Heatmap for {protein_pocket_base}", pad=20)
    plt.ylabel("")
    plt.xlabel("")
    plt.xticks(rotation=45, ha='right', rotation_mode='anchor', fontsize=8)
    plt.yticks(rotation=0)
    plt.tight_layout()
    
    fig.savefig(str(tiff_path), format="tiff", dpi=600, bbox_inches="tight")
    plt.close(fig)
    
    log_step(None, f"Saved heatmap for {protein_pocket_base}", color="white")

def print_ranking(results, output_csv=None, ligand_names=None, protein_names=None):
    if not results:
        logging.warning("No valid log files or energy scores found.")
        return []
        
    import statistics

    energies_dict = {}
    for protein, pocket, ligand, run, energy in results:
        key = (protein, pocket, ligand)
        if key not in energies_dict:
            energies_dict[key] = []
        energies_dict[key].append(energy)
        
    mean_energies = {}
    for key, energies in energies_dict.items():
        mean_energies[key] = statistics.mean(energies)

    best_dict = {}
    for protein, pocket, ligand, run, energy in results:
        base_ligand = _get_base_ligand(ligand)
        key = (protein, pocket, base_ligand)
        isomer_key = (protein, pocket, ligand)
        isomer_mean = mean_energies[isomer_key]
        
        # Keep the isomer with the lowest mean energy
        if key not in best_dict:
            best_dict[key] = (isomer_mean, (protein, pocket, ligand, run, energy))
        else:
            if isomer_mean < best_dict[key][0]:
                best_dict[key] = (isomer_mean, (protein, pocket, ligand, run, energy))
            elif isomer_mean == best_dict[key][0]:
                # Tie-breaker: lowest single energy
                if energy < best_dict[key][1][4]:
                    best_dict[key] = (isomer_mean, (protein, pocket, ligand, run, energy))
            
    best_results = [val[1] for val in best_dict.values()]

    curated_results_with_stats = []
    
    for row in best_results:
        protein, pocket, ligand, run, energy = row
        key = (protein, pocket, ligand)
        energies = energies_dict.get(key, [energy])
        mean_e = statistics.mean(energies)
        sd_e = statistics.stdev(energies) if len(energies) > 1 else None
        
        curated_results_with_stats.append((protein, pocket, ligand, run, energy, mean_e, sd_e))

    # Sort curated results by mean affinity (lowest/most negative first)
    curated_results_with_stats.sort(key=lambda x: x[5])

    # Load mappings
    cid_to_name = _load_mapping_from_csv(ligand_names)
    protein_to_name = _load_mapping_from_csv(protein_names)

    def format_row(row):
        protein, pocket, ligand, run, energy = row
        cid = _get_base_ligand(ligand)
        if "_isomer_" in ligand:
            parts = ligand.split("_isomer_")
            isomer = parts[1] if len(parts) > 1 else ""
        else:
            isomer = "N/A"
        ligand_name = cid_to_name.get(cid.lower(), "N/A")
        protein_name = protein_to_name.get(protein.lower(), "N/A")
        
        base_res = [protein, protein_name, pocket, cid, ligand_name, isomer, run, energy]
        return base_res

    header_raw = ['Protein', 'Protein_name', 'Pocket ID', 'CID', 'Ligand_name', 'Isomer', 'Run', 'Affinity (kcal/mol)']
    header_curated = ['Protein', 'Protein_name', 'Pocket ID', 'CID', 'Ligand_name', 'Isomer', 'Run', 'Lowest affinity (kcal/mol)', 'Mean affinity (kcal/mol)', 'SD']

    formatted_results = [format_row(row) for row in results]
    
    formatted_curated = []
    for cur_res in curated_results_with_stats:
        protein, pocket, ligand, run, energy, mean_e, sd_e = cur_res
        base_row = format_row((protein, pocket, ligand, run, energy))
        
        mean_str = f"{mean_e:.2f}".replace("-", "\u2212")
        sd_str = f"{sd_e:.2f}" if sd_e is not None else "N/A"
        
        base_row.extend([mean_str, sd_str])
        formatted_curated.append(base_row)

    print()
    display_limit = 10
    log_step(None, f"--- Top {min(display_limit, len(formatted_curated))} Curated Best Complexes by Mean Free Energy (Total: {len(formatted_curated)}) ---", color="magenta")
    log_step(None, f"{'Protein':<15} | {'Pocket ID':<10} | {'CID':<15} | {'Lowest affinity (kcal/mol)':<28} | {'Mean affinity (kcal/mol)':<26} | {'SD':<10}", color="magenta")
    log_step(None, "-" * 115, color="magenta")
    for row in formatted_curated[:display_limit]:
        protein, protein_name, pocket, cid, lname, iso, run, energy, mean_str, sd_str = row[:10]
        log_step(None, f"{protein:<15} | {pocket:<10} | {cid:<15} | {energy:<28.2f} | {mean_str:<26} | {sd_str:<10}", color="magenta")

    if output_csv:
        try:
            with open(output_csv, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(header_raw)
                writer.writerows(formatted_results)
            print()
            log_step(None, f"Raw ranking saved to {output_csv}", color="magenta")
            
            best_csv = Path(output_csv).with_name(f"curated_best_{Path(output_csv).name}")
            with open(best_csv, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(header_curated)
                writer.writerows(formatted_curated)
            log_step(None, f"Curated best ranking saved to {best_csv}", color="magenta")
                
        except Exception as e:
            logging.error(f"Error saving to CSV: {e}")

    return curated_results_with_stats

def _convert_to_pdb(input_file, in_format, output_file, extra_args=None):
    import subprocess
    cmd = ["obabel", "-i", in_format, str(input_file)]
    if extra_args:
        cmd.extend(extra_args)
    cmd.extend(["-o", "pdb", "-O", str(output_file)])
    subprocess.run(cmd, check=True, capture_output=True)

def _find_available_chain(pdb_file):
    import string
    used_chains = set()
    if pdb_file.exists():
        with open(pdb_file, 'r') as p_f:
            for line in p_f:
                if line.startswith(("ATOM", "HETATM")) and len(line) > 21:
                    used_chains.add(line[21])
    
    # Legacy PDB allows A-Z, a-z, and 0-9
    all_candidates = "ZYXWVUTSRQPONMLKJIHGFEDCBA" + string.ascii_lowercase[::-1] + string.digits[::-1]
    for candidate in all_candidates:
        if candidate not in used_chains:
            return candidate
    return None

def _combine_complex_pdbs(prot_pdb, lig_pdb, complex_pdb, lig_chain):
    max_atom_serial = 0
    with open(complex_pdb, 'w') as out_f:
        # Write protein, skipping END and CONECT lines
        with open(prot_pdb, 'r') as p_f:
            for line in p_f:
                if not line.startswith(("END", "CONECT", "MASTER")):
                    if line.startswith(("ATOM", "HETATM")) and len(line) >= 11:
                        try:
                            serial = int(line[6:11].strip())
                            if serial > max_atom_serial:
                                max_atom_serial = serial
                        except ValueError:
                            pass
                    out_f.write(line)
        # Write ligand, skipping CONECT lines and adding a chain ID
        with open(lig_pdb, 'r') as l_f:
            for line in l_f:
                if not line.startswith(("END", "CONECT", "MASTER", "COMPND", "AUTHOR")):
                    # Assign the unique chain to the ligand to differentiate it from the protein
                    if line.startswith(("ATOM", "HETATM")) and len(line) > 21:
                        line = line[:21] + lig_chain + line[22:]
                        
                        # Renumber atom serial
                        max_atom_serial += 1
                        serial_str = f"{max_atom_serial:>5}"
                        # If it exceeds 5 characters (e.g. 100,000), typical PDB format overflows.
                        if len(serial_str) > 5:
                            raise ValueError(f"Atom serial number {max_atom_serial} exceeds 99999 (legacy PDB limit).")
                        line = line[:6] + serial_str + line[11:]
                        
                    out_f.write(line)
        # Write END
        out_f.write("END\n")

def _map_protein_chains(tmp_prot_pdb, mapped_prot_pdb, protein, asym_id_mapping):
    import logging
    import string
    
    protein_mapping = asym_id_mapping.get(protein) or asym_id_mapping.get(f"{protein}.cif") or {}
    prody_to_orig = {v: k for k, v in protein_mapping.items()}
    
    existing_chains = set()
    with open(tmp_prot_pdb, 'r') as in_f:
        for line in in_f:
            if line.startswith(("ATOM", "HETATM")) and len(line) > 21:
                existing_chains.add(line[21])
                
    used_chains = set()
    for current_chain, orig_chain in prody_to_orig.items():
        if len(orig_chain) == 1:
            used_chains.add(orig_chain)
            
    for current_chain in existing_chains:
        if current_chain not in prody_to_orig:
            used_chains.add(current_chain)
            
    all_candidates = "ZYXWVUTSRQPONMLKJIHGFEDCBA" + string.ascii_lowercase[::-1] + string.digits[::-1]
    
    final_mapping = {}
    for current_chain, orig_chain in prody_to_orig.items():
        if len(orig_chain) == 1:
            final_mapping[current_chain] = orig_chain
        else:
            assigned = False
            for candidate in all_candidates:
                if candidate not in used_chains:
                    final_mapping[current_chain] = candidate
                    used_chains.add(candidate)
                    logging.warning(f"Original chain ID '{orig_chain}' for {protein} is > 1 char. Reassigned to '{candidate}'.")
                    assigned = True
                    break
            if not assigned:
                logging.warning(f"Cannot reassign chain for '{orig_chain}' in {protein}. Legacy PDB limit reached.")
                return False, None
    
    with open(tmp_prot_pdb, 'r') as in_f, open(mapped_prot_pdb, 'w') as out_f:
        for line in in_f:
            if line.startswith(("ATOM", "HETATM")) and len(line) > 21:
                current_chain = line[21]
                if current_chain in final_mapping:
                    line = line[:21] + final_mapping[current_chain] + line[22:]
            out_f.write(line)
            
    orig_to_final = {orig: final_mapping[curr] for curr, orig in prody_to_orig.items() if curr in final_mapping}
    return True, orig_to_final

def generate_complexes(results, output_dir, protein_clean_dir, display_limit=20):
    if not results:
        return
    
    import subprocess
    import tempfile
    import logging
    import os
    import json
    from pathlib import Path
    from logger_utils import log_step

    vis_dir = Path(output_dir) / "visualization"
    os.makedirs(vis_dir, exist_ok=True)
    
    mapping_file = Path(protein_clean_dir) / "asym_id_mapping.json"
    asym_id_mapping = {}
    if mapping_file.exists():
        try:
            with open(mapping_file, "r") as f:
                asym_id_mapping = json.load(f)
        except json.JSONDecodeError:
            logging.warning(f"Failed to parse {mapping_file}")

    log_step(None, f"Generating PDB complex files for top {min(display_limit, len(results))} results...", color="white")
    
    complex_chain_mappings = {}
    mapping_out_file = vis_dir / "final_chain_mapping.json"
    if mapping_out_file.exists():
        try:
            with open(mapping_out_file, "r") as f:
                complex_chain_mappings = json.load(f)
        except Exception:
            pass

    for row in results[:display_limit]:
        protein, pocket, ligand, run, energy = row[:5]
        protein_pocket_base = f"{protein}_pocket_{pocket}" if pocket != "N/A" else protein
        
        # Paths
        protein_cif = Path(protein_clean_dir) / f"{protein}FH.cif"
        
        ligand_sdf = Path(output_dir) / "vina_output" / f"run_{run}" / f"{protein_pocket_base}_{ligand}" / f"{protein_pocket_base}_{ligand}_vina_out.sdf"
        
        complex_pdb = vis_dir / f"{protein_pocket_base}_{ligand}_complex.pdb"
        
        if complex_pdb.exists():
            logging.info(f"Complex file {complex_pdb.name} already exists. Skipping.")
            continue

        if not protein_cif.exists():
            logging.warning(f"Protein file missing: {protein_cif}. Skipping complex generation.")
            continue
            
        if not ligand_sdf.exists():
            logging.warning(f"Ligand SDF file missing: {ligand_sdf}. Skipping complex generation.")
            continue
            
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_prot_pdb = Path(tmpdir) / "prot.pdb"
            tmp_lig_pdb = Path(tmpdir) / "lig.pdb"
            
            # Convert CIF to PDB
            try:
                _convert_to_pdb(protein_cif, "cif", tmp_prot_pdb)
            except subprocess.CalledProcessError as e:
                logging.error(f"Failed to convert protein CIF to PDB: {e}")
                continue
                
            # Convert SDF to PDB
            try:
                _convert_to_pdb(ligand_sdf, "sdf", tmp_lig_pdb, extra_args=["-f", "1", "-l", "1"])
            except subprocess.CalledProcessError as e:
                logging.error(f"Failed to convert ligand SDF to PDB: {e}")
                continue
                
            # Apply original chain ID mapping
            mapped_prot_pdb = Path(tmpdir) / "mapped_prot.pdb"
            success, orig_to_final = _map_protein_chains(tmp_prot_pdb, mapped_prot_pdb, protein, asym_id_mapping)
            if not success:
                logging.warning(f"Skipping complex for {protein} due to legacy PDB chain limitation.")
                continue

            # Determine available chain ID for ligand
            lig_chain = _find_available_chain(mapped_prot_pdb)
            if lig_chain is None:
                logging.warning(f"No available chain IDs left for {protein_pocket_base}_{ligand} (Legacy PDB limit reached). Skipping complex generation.")
                continue

            # Combine PDBs
            try:
                _combine_complex_pdbs(mapped_prot_pdb, tmp_lig_pdb, complex_pdb, lig_chain)
                logging.debug(f"Created complex: {complex_pdb}")
                orig_to_final["LIGAND"] = lig_chain
                complex_chain_mappings[f"{protein_pocket_base}_{ligand}"] = orig_to_final
            except Exception as e:
                logging.warning(f"Skipping complex {protein_pocket_base}_{ligand}: {e}")
                if complex_pdb.exists():
                    try:
                        complex_pdb.unlink()
                    except OSError:
                        pass
                
    if complex_chain_mappings:
        mapping_out_file = vis_dir / "final_chain_mapping.json"
        try:
            with open(mapping_out_file, "w") as f:
                json.dump(complex_chain_mappings, f, indent=4)
            log_step(None, f"Final chain mappings saved to {mapping_out_file}", color="white")
        except Exception as e:
            logging.error(f"Failed to save final chain mappings: {e}")
            
    log_step(None, f"Complexes saved to {vis_dir}", color="white")

def visualize_prolif_results(results, output_dir, protein_clean_dir, ligand_path=None, display_limit=20, ligand_names=None, protein_names=None):
    if not results:
        return
        
    import os
    import logging
    import pandas as pd
    import matplotlib.pyplot as plt
    from pathlib import Path
    from logger_utils import log_step
    import warnings
    import sys
    
    warnings.simplefilter("ignore")
    os.environ["PYTHONWARNINGS"] = "ignore"
    
    devnull = open(os.devnull, 'w')
    old_stderr = sys.stderr
    sys.stderr = devnull
    try:
        import MDAnalysis as mda
        import prolif as plf
        from prolif.plotting.barcode import Barcode
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        import seaborn as sns
    except ImportError as e:
        sys.stderr = old_stderr
        devnull.close()
        logging.error(f"Missing dependency for ProLif visualization: {e}")
        log_step(None, f"Skipping ProLif visualization due to missing dependency: {e}", color="yellow")
        return
    finally:
        sys.stderr = old_stderr
        devnull.close()

    vis_dir = Path(output_dir) / "visualization" / "prolif"
    os.makedirs(vis_dir / "minimal", exist_ok=True)
    os.makedirs(vis_dir / "all", exist_ok=True)
    
    log_step(None, f"Generating ProLif visualizations for top {min(display_limit, len(results))} results...", color="white")

    # Setup headless chrome
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--window-size=1200,800")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    
    try:
        driver = webdriver.Chrome(options=chrome_options)
        
        # Force a high-resolution render (Device Scale Factor)
        metrics = {
            "width": 1200,
            "height": 800,
            "deviceScaleFactor": 3,
            "mobile": False
        }
        driver.execute_cdp_cmd("Emulation.setDeviceMetricsOverride", metrics)
    except Exception as e:
        logging.error(f"Failed to initialize Selenium Chrome driver: {e}")
        log_step(None, "Skipping ProLif visualization because Selenium Chrome driver could not be initialized.", color="yellow")
        return

    # Map CIDs and Proteins to standard names
    cid_to_name = _load_mapping_from_csv(ligand_names)
    protein_to_name = _load_mapping_from_csv(protein_names)
    protein_pockets = _get_protein_pockets(results)

    # Set publication typography globally
    plt.rcParams.update({
        'font.family': 'sans-serif',
        'font.sans-serif': ['Arial', 'Helvetica', 'DejaVu Sans'],
        'font.size': 12,
        'axes.labelsize': 14,
        'xtick.labelsize': 12,
        'ytick.labelsize': 10,
        'legend.fontsize': 10
    })
    Barcode.COLORS["Hydrophobic"] = "#4477AA"
    Barcode.COLORS["HBAcceptor"] = "#EE6677"
    Barcode.COLORS["HBDonor"] = "#228833"
    Barcode.COLORS["PiStacking"] = "#CCBB44"

    # Group results by Protein and Pocket
    targets = {}
    for row in results[:display_limit]:
        protein, pocket, ligand, run, energy, mean_e, sd_e = row
        key = (protein, pocket)
        if key not in targets:
            targets[key] = []
        targets[key].append(row)

    try:
        all_interactions = plf.Fingerprint.list_available()
        
        excel_data_rows = []
        complex_no = 1
        
        for (protein, pocket), target_results in targets.items():
            protein_pocket_base = f"{protein}_pocket_{pocket}" if pocket != "N/A" else protein
            display_protein_pocket = _format_display_protein_pocket(protein, pocket, protein_to_name, protein_pockets)
                
            protein_cif = Path(protein_clean_dir) / f"{protein}FH.cif"
            
            if not protein_cif.exists():
                logging.warning(f"Missing {protein_cif}. Skipping visualization for {protein_pocket_base}.")
                continue

            fp_dataframes_minimal = []
            ligand_labels_minimal = []
            
            fp_dataframes_all = []
            ligand_labels_all = []

            try:
                with _prepare_prolif_protein(protein_cif) as protein_mol:
                    for row in target_results:
                        _, _, ligand, run, energy, mean_e, sd_e = row
                        ligand_sdf = Path(output_dir) / "vina_output" / f"run_{run}" / f"{protein_pocket_base}_{ligand}" / f"{protein_pocket_base}_{ligand}_vina_out.sdf"
                        
                        if not ligand_sdf.exists():
                            continue

                        base_ligand = _get_base_ligand(ligand)
                        label = cid_to_name.get(base_ligand.lower(), base_ligand)
                        
                        poses = list(plf.sdf_supplier(str(ligand_sdf)))
                        if not poses:
                            continue
                            
                        versions = [
                            ("minimal", {}, fp_dataframes_minimal, ligand_labels_minimal),
                            ("all", {"count": True}, fp_dataframes_all, ligand_labels_all)
                        ]
                        
                        for v_name, fp_kwargs, fp_dfs, lig_labels in versions:
                            fp = plf.Fingerprint(all_interactions, **fp_kwargs)
                            fp.run_from_iterable([poses[0]], protein_mol, progress=False)
                            
                            out_png = vis_dir / v_name / f"{display_protein_pocket}_{ligand}_prolif.png"
                            if out_png.exists():
                                logging.info(f"ProLif plot {out_png.name} already exists. Skipping.")
                            else:
                                try:
                                    _generate_network_plot(fp, poses[0], driver, out_png)
                                except Exception as e:
                                    logging.error(f"Failed to generate {v_name} network plot for {protein_pocket_base}_{ligand}: {e}")
                            
                            df = fp.to_dataframe()
                            
                            if 'ligand' in df.columns.names:
                                idx = df.columns.names.index('ligand')
                                new_tuples = [tuple('LIG' if i == idx else val for i, val in enumerate(tup)) for tup in df.columns]
                                df.columns = pd.MultiIndex.from_tuples(new_tuples, names=df.columns.names)
                                
                            fp_dfs.append(df)
                            lig_labels.append(label)

                            if v_name == "all":
                                interaction_groups = {}
                                if not df.empty:
                                    row_data = df.iloc[0]
                                    for col, val in row_data.items():
                                        if val > 0:
                                            lig_res, prot_res, interaction = col
                                            if interaction not in interaction_groups:
                                                interaction_groups[interaction] = []
                                            interaction_groups[interaction].append(f"{prot_res} ({int(val)})")
                                
                                formatted_ligand = label
                                energy_val = float(f"{mean_e:.3f}")
                                
                                if not interaction_groups:
                                    excel_data_rows.append({
                                        'complex_id': complex_no,
                                        'No': complex_no,
                                        'Protein': display_protein_pocket,
                                        'Ligand': formatted_ligand,
                                        'Docking score (kcal/mol)': energy_val,
                                        'Interaction': "None",
                                        'Amino Acid (number of interaction)': "None"
                                    })
                                else:
                                    for interaction, amino_acids in interaction_groups.items():
                                        for aa in amino_acids:
                                            excel_data_rows.append({
                                                'complex_id': complex_no,
                                                'No': complex_no,
                                                'Protein': display_protein_pocket,
                                                'Ligand': formatted_ligand,
                                                'Docking score (kcal/mol)': energy_val,
                                                'Interaction': interaction,
                                                'Amino Acid (number of interaction)': aa
                                            })
                                complex_no += 1
                        
            except Exception as e:
                logging.error(f"Failed to process protein {protein} or its ligands for ProLif visualization: {e}")
                continue

            for fp_dfs, lbls, version in [(fp_dataframes_minimal, ligand_labels_minimal, "minimal"),
                                          (fp_dataframes_all, ligand_labels_all, "all")]:
                if fp_dfs:
                    df_combined = pd.concat(fp_dfs, ignore_index=True)
                    if version == "all":
                        df_combined = df_combined.fillna(0)
                        df_combined.index = lbls
                        
                        tiff_path = vis_dir / version / f"{display_protein_pocket}_heatmap.tiff"
                        try:
                            _generate_heatmap_plot(df_combined, tiff_path, display_protein_pocket)
                        except Exception as e:
                            logging.error(f"Failed to generate {version} heatmap plot for {protein_pocket_base}: {e}")
                    else:
                        df_combined = df_combined.fillna(False)
                        df_combined.index = lbls
                        
                        tiff_path = vis_dir / version / f"{display_protein_pocket}_barcode.tiff"
                        try:
                            _generate_barcode_plot(df_combined, tiff_path, display_protein_pocket)
                        except Exception as e:
                            logging.error(f"Failed to generate {version} barcode plot for {protein_pocket_base}: {e}")

        # Generate Excel report
        if excel_data_rows:
            _generate_excel_summary(excel_data_rows, vis_dir)

        log_step(None, f"All ProLif visualizations saved to {vis_dir}", color="magenta")
    finally:
        driver.quit()

def _generate_excel_summary(excel_data_rows, vis_dir):
    import logging
    import pandas as pd
    from logger_utils import log_step
    
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side
        
        excel_df = pd.DataFrame(excel_data_rows)
        
        # Sort by Docking score
        excel_df = excel_df.sort_values(by=['Docking score (kcal/mol)', 'complex_id'])
        
        excel_path = vis_dir / "prolif_interactions_summary.xlsx"
        
        # Write to Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Interactions"
        
        columns = ['Protein', 'Ligand', 'Docking score (kcal/mol)', 'Interaction', 'Amino Acid (number of interaction)']
        ws.append(columns)
        
        # Style header
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Write data
        for idx, row in excel_df.iterrows():
            ws.append([row[col] for col in columns])
        
        # Apply borders and alignment to all cells
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Merge cells
        def merge_column_by_key(col_letter, key_col_name, merge_by_interaction=False):
            start_row = 2
            while start_row <= ws.max_row:
                key_val = excel_df.iloc[start_row - 2][key_col_name]
                if merge_by_interaction:
                    interaction_val = excel_df.iloc[start_row - 2]['Interaction']
                
                end_row = start_row
                while end_row < ws.max_row:
                    next_key_val = excel_df.iloc[end_row - 1][key_col_name]
                    if next_key_val != key_val:
                        break
                    if merge_by_interaction:
                        next_interaction_val = excel_df.iloc[end_row - 1]['Interaction']
                        if next_interaction_val != interaction_val:
                            break
                    end_row += 1
                
                if end_row > start_row:
                    ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                start_row = end_row + 1

        merge_column_by_key('A', 'complex_id')
        merge_column_by_key('B', 'complex_id')
        merge_column_by_key('C', 'complex_id')
        merge_column_by_key('D', 'complex_id', merge_by_interaction=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        
        wb.save(str(excel_path))
        log_step(None, f"Excel interaction summary saved to {excel_path}", color="white")
    except ImportError:
        logging.warning("openpyxl is not installed. Skipping Excel summary generation. (pip install openpyxl)")
    except Exception as e:
        logging.error(f"Failed to generate Excel summary: {e}")

def generate_ranking_heatmap(curated_results, output_dir, ligand_names=None, protein_names=None, display_limit=20, positive_control_map=None):
    """
    Generates a heatmap of docking scores (affinity) for top compounds across different protein pockets.
    X-axis: Protein_pocket
    Y-axis: Ligand names (top N based on overall mean energy)
    """
    if not curated_results:
        return

    from pathlib import Path
    import logging
    from logger_utils import log_step

    try:
        import pandas as pd
        import matplotlib.pyplot as plt
        import seaborn as sns
        
        # Load mappings
        cid_to_name = _load_mapping_from_csv(ligand_names)
        protein_to_name = _load_mapping_from_csv(protein_names)
        protein_pockets = _get_protein_pockets(curated_results)
        
        data = []
        for protein, pocket, ligand, run, energy, mean_e, sd_e in curated_results:
            protein_pocket = _format_display_protein_pocket(protein, pocket, protein_to_name, protein_pockets)
                
            base_ligand = _get_base_ligand(ligand)
            is_positive_control = False
            if positive_control_map and base_ligand in positive_control_map:
                if protein in positive_control_map[base_ligand]:
                    is_positive_control = True
                    
            if is_positive_control:
                data.append({
                    "Protein_pocket": protein_pocket,
                    "Ligand": "Positive Control",
                    "Energy": mean_e
                })
            else:
                # Address user's feedback: _get_base_ligand retrieves the base string from curated_results.
                # Then map to common name if it exists, otherwise keep base_ligand.
                ligand_name = cid_to_name.get(base_ligand.lower(), base_ligand)
                data.append({
                    "Protein_pocket": protein_pocket,
                    "Ligand": ligand_name,
                    "Energy": mean_e
                })
            
        df = pd.DataFrame(data)
        
        if df.empty:
            logging.warning("No data available to generate ranking heatmap.")
            return

        # Filter out positive scores (>= 0) for ranking calculations so poor binding doesn't heavily penalize the mean
        df_for_ranking = df[df['Energy'] < 0]
        
        # Identify top N ligands based on their best (mean) overall docking score across all proteins
        top_ligands = df_for_ranking.groupby('Ligand')['Energy'].mean().nsmallest(display_limit).index.tolist()
        if "Positive Control" not in top_ligands:
            top_ligands.append("Positive Control")
        
        # Identify top N proteins/pockets based on their best (mean) overall docking score
        top_proteins = df_for_ranking.groupby('Protein_pocket')['Energy'].mean().nsmallest(display_limit).index
        
        # Filter dataframe for only those top ligands AND top proteins
        df_filtered = df[df['Ligand'].isin(top_ligands) & df['Protein_pocket'].isin(top_proteins)]
        
        # Pivot the dataframe to create a matrix for the heatmap
        heatmap_data = df_filtered.pivot_table(index='Ligand', columns='Protein_pocket', values='Energy', aggfunc='mean')
        
        # Replace positive docking scores (poor binding) with NaN to avoid skewing color scale
        import numpy as np
        heatmap_data = heatmap_data.mask(heatmap_data >= 0, np.nan)
        
        # Sort Y-axis (Ligands) by the overall best energy (excluding >= 0 scores)
        sorted_ligands = df_filtered[df_filtered['Energy'] < 0].groupby('Ligand')['Energy'].mean().sort_values().index.tolist()
        
        # Ensure 'Positive Control' is always at the top
        if "Positive Control" in sorted_ligands:
            sorted_ligands.remove("Positive Control")
            sorted_ligands.insert(0, "Positive Control")
            
        heatmap_data = heatmap_data.loc[sorted_ligands]
        
        # Sort X-axis (Proteins) by the overall best energy (excluding >= 0 scores)
        sorted_proteins = df_filtered[df_filtered['Energy'] < 0].groupby('Protein_pocket')['Energy'].mean().sort_values().index
        heatmap_data = heatmap_data[sorted_proteins]
        
        # Plot
        fig_width = max(10, len(heatmap_data.columns) * 1.5)
        fig_height = max(8, len(heatmap_data.index) * 0.5)
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        
        # We use a colormap where lower values (better affinity) are distinct.
        sns.heatmap(heatmap_data, cmap='YlGnBu_r', annot=True, fmt=".1f", 
                    cbar_kws={'label': 'Docking Score (kcal/mol)'},
                    linewidths=.5, ax=ax, 
                    mask=heatmap_data.isnull())  # Missing data is masked
        
        # Invert the colorbar so dark color (lowest score) is at the top
        cbar = ax.collections[0].colorbar
        cbar.ax.invert_yaxis()
        
        plt.title(f"Docking Scores Heatmap (Top {len(sorted_ligands)} Compounds)", pad=20, fontsize=14)
        plt.xlabel("Protein / Pocket", fontsize=12)
        plt.ylabel("Compound", fontsize=12)
        plt.xticks(rotation=45, ha='right', fontsize=10)
        plt.yticks(rotation=0, fontsize=10)
        plt.tight_layout()
        
        vis_dir = Path(output_dir) / "visualization"
        vis_dir.mkdir(parents=True, exist_ok=True)
        heatmap_path = vis_dir / "ranking_heatmap.tiff"
        
        fig.savefig(str(heatmap_path), format="tiff", dpi=600, bbox_inches="tight")
        plt.close(fig)
        
        log_step(None, f"Ranking heatmap saved to {heatmap_path}", color="magenta")

    except Exception as e:
        logging.error(f"Failed to generate ranking heatmap: {e}")